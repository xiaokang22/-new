# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO

def export_excel(records, title):
    """导出Excel报表"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "业绩报表"
    
    # 定义样式
    title_font = Font(name='微软雅黑', size=16, bold=True)
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    normal_font = Font(name='微软雅黑', size=10)
    number_format = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题
    ws.merge_cells('A1:F1')
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # 空行
    ws.append([])
    
    # 表头
    headers = ['日期', '渠道', '业务员', '金额', '备注', '创建时间']
    ws.append(headers)
    
    # 设置表头样式
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # 数据行
    for record in records:
        row_data = [
            record.get('date', ''),
            '到店购买' if record.get('channel') == 'store' else '业务员推销',
            record.get('salesperson_name', '-'),
            record.get('amount', 0),
            record.get('note', ''),
            record.get('created_at', '')
        ]
        ws.append(row_data)
        
        # 设置数据行样式
        row_num = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = normal_font
            cell.border = thin_border
            if col == 4:  # 金额列
                cell.number_format = number_format
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='center')
    
    # 汇总行
    ws.append([])
    ws.append(['汇总', '', '', f'=SUM(D4:D{ws.max_row})', '', ''])
    summary_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=summary_row, column=col)
        cell.font = Font(name='微软雅黑', size=10, bold=True)
        cell.border = thin_border
    
    # 设置列宽
    column_widths = [15, 15, 15, 15, 25, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
